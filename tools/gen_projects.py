"""Regenerate all per-project HTML files + the projects.html cards grid.

Reads:
  - assets/BA Holding.xlsx        (master data, sheets 1..13)
  - tools/template-project-item.html (frozen NMZ-style layout used as the template)

Writes:
  - project-{slug}.html  (one per project, 13 total)
  - projects.html        (updates the <div class="proj-grid" id="projGrid">…</div>
                          block in place, leaves everything else untouched)

Run from the repo root:

    python tools/gen_projects.py

Requires the openpyxl package (pip install openpyxl).
"""

from __future__ import annotations

import io
import json
import re
import sys
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'assets' / 'BA Holding.xlsx'
TEMPLATE = ROOT / 'tools' / 'template-project-item.html'
PROJECTS_HTML = ROOT / 'projects.html'

# ---------- 1. Excel → raw dict per sheet ----------
KEYS = {
    'Направление':       'direction',
    'Название':          'name',
    'Описание':          'desc',
    'Годовой оборот':    'turnover',
    'Сотрудники':        'employees',
    'Мощность в год':    'capacity',
    'Год запуска':       'year',
    'О предприятии: Заголовок':      'about_h',
    'О предприятии: Полное описание':'about_full',
    'О предприятии: Продукты (макисмум 1-4)': 'products',
    'Gmail':    'email',
    'Телефон':  'phone',
    'Сайт':     'site',
    'Локация':  'location',
}

def load_excel():
    wb = openpyxl.load_workbook(str(XLSX), data_only=True)
    out = {}
    for name in wb.sheetnames:
        if name.startswith('BA Holding'):
            continue
        ws = wb[name]
        info = {}
        for row in ws.iter_rows(values_only=True):
            if not row or not row[0]:
                continue
            k = str(row[0]).strip().rstrip(':').strip()
            if k in KEYS:
                v = row[1] if len(row) > 1 else None
                if v is None:
                    continue
                if isinstance(v, float) and v.is_integer():
                    v = int(v)
                info[KEYS[k]] = str(v).strip() if not isinstance(v, int) else v
        if info.get('name'):
            out[name] = info
    return out

# ---------- 2. Per-project metadata (slug / category / logo / image / hero text) ----------
META = {
    '1. NMZ Namangan Metallurgiya za': dict(
        slug='nmz', cat='metallurgy', catLabel='Металлургия',
        logo='assets/img/logos/NMZ-logo 26.webp', image=1,
        shortLoc='Наманганская обл., г. Чуст',
        siteLabel='nmz.uz', siteUrl='https://nmz.uz/',
        breadcrumb='Namangan Metallurgiya Zavodi',
        h1_l1='Namangan Metallurgiya Zavodi', h1_l2='(NMZ)',
        overviewH1='Производственная мощь', overviewH2='Ферганской долины',
        relatedHeading='металлургические активы',
        banner='assets/img/projects/other/other (1).webp',
        gallery=[
            'assets/img/projects/other/other (2).webp',
            'assets/img/projects/other/other (3).webp',
            'assets/img/projects/other/other (4).webp',
            'assets/img/projects/other/other (5).webp',
        ],
    ),
    '2. TMZ Tojikiston metallurgiya ': dict(
        slug='tmz', cat='metallurgy', catLabel='Металлургия',
        logo='assets/img/logos/TMZ.webp', image=4,
        shortLoc='Таджикистан, Согд, Гулистон',
        breadcrumb='Таджикский металлургический завод',
        h1_l1='Таджикский', h1_l2='металлургический завод (TMZ)',
        overviewH1='Металлургия для', overviewH2='Центральной Азии',
        relatedHeading='металлургические активы',
        banner='assets/img/projects/tmz/banner.webp',
        gallery=[
            'assets/img/projects/tmz/1.webp',
            'assets/img/projects/tmz/2.webp',
            'assets/img/projects/tmz/3.webp',
            'assets/img/projects/tmz/4.webp',
        ],
    ),
    '3. JMZ Jizzax metallurgiya zavo': dict(
        slug='jmz', cat='metallurgy', catLabel='Металлургия',
        logo='assets/img/logos/JMZ-logo (horizontal dark).webp', image=2,
        shortLoc='Джизакская обл., г. Джизак',
        siteLabel='jmz.uz', siteUrl='https://jmz.uz/',
        breadcrumb='Джизакский металлургический завод',
        h1_l1='Джизакский', h1_l2='металлургический завод (JMZ)',
        overviewH1='Современное', overviewH2='металлургическое производство',
        relatedHeading='металлургические активы',
        banner='assets/img/projects/jmz/banner.webp',
        gallery=[
            'assets/img/projects/jmz/1.webp',
            'assets/img/projects/jmz/2.webp',
            'assets/img/projects/jmz/3.webp',
            'assets/img/projects/jmz/4.webp',
            'assets/img/projects/jmz/5.webp',
            'assets/img/projects/jmz/6.webp',
        ],
    ),
    '4. BA Chirchiq Metall': dict(
        slug='ba-chirchiq-metall', cat='metallurgy', catLabel='Металлургия',
        logo='assets/img/logos/BA Chiqchiq metall logo 26.webp', image=5,
        shortLoc='Ташкентская обл., г. Чирчик',
        breadcrumb='BA Chirchiq Metall',
        h1_l1='BA Chirchiq Metall', h1_l2='Чирчикский металлургический завод',
        overviewH1='Стальные шары для', overviewH2='промышленности и строительства',
        relatedHeading='металлургические активы',
        banner='assets/img/projects/chirchik-metall/banner.webp',
        gallery=[
            'assets/img/projects/chirchik-metall/1.webp',
            'assets/img/projects/chirchik-metall/2.webp',
            'assets/img/projects/chirchik-metall/3.webp',
        ],
    ),
    '5. Emirate Steel': dict(
        slug='emirate-steel', cat='metallurgy', catLabel='Металлургия',
        logo='assets/img/logos/Emirate steel 26-3.webp', image=3,
        shortLoc='Джизакская обл., Бахмальский р-н',
        breadcrumb='Emirate Steel',
        h1_l1='Emirate Steel', h1_l2='Металлопрокат для строительства',
        overviewH1='Современный завод', overviewH2='металлопроката',
        relatedHeading='металлургические активы',
        banner='assets/img/projects/other/other (6).webp',
        gallery=[
            'assets/img/projects/other/other (7).webp',
            'assets/img/projects/other/other (8).webp',
            'assets/img/projects/other/other (9).webp',
            'assets/img/projects/other/other (10).webp',
        ],
    ),
    '6. Fu-xin': dict(
        slug='fuxin', cat='production', catLabel='Производство',
        logo='assets/img/logos/Фухин.webp', image=6,
        shortLoc='Ташкент, Бектемирский р-н',
        siteLabel='fuxin.uz', siteUrl='https://fuxin.uz/',
        breadcrumb='Fu-xin',
        h1_l1='Fu-xin', h1_l2='Металлоконструкции и горячее цинкование',
        overviewH1='Металлоконструкции и', overviewH2='горячее цинкование',
        relatedHeading='производственные активы',
        banner='assets/img/projects/fuxin/banner.webp',
        gallery=[
            'assets/img/projects/fuxin/1.webp',
            'assets/img/projects/fuxin/2.webp',
            'assets/img/projects/fuxin/3.webp',
            'assets/img/projects/fuxin/4.webp',
            'assets/img/projects/fuxin/5.webp',
            'assets/img/projects/fuxin/6.webp',
        ],
    ),
    '7. China House ': dict(
        slug='china-house', cat='construction', catLabel='Строительство',
        logo='assets/img/logos/China house-logo 26.webp', image=8,
        shortLoc='Ташкент, Бектемирский р-н',
        breadcrumb='China House',
        h1_l1='China House', h1_l2='Жилые и коммерческие комплексы',
        overviewH1='Строительство', overviewH2='современной недвижимости',
        relatedHeading='строительные проекты',
        banner='assets/img/projects/china-house/banner.webp',
        gallery=[
            'assets/img/projects/china-house/1.webp',
            'assets/img/projects/china-house/2.webp',
            'assets/img/projects/china-house/3.webp',
            'assets/img/projects/china-house/5.webp',
            'assets/img/projects/china-house/6.webp',
            'assets/img/projects/china-house/7.webp',
        ],
    ),
    '8. Toshkent Plast Polimer': dict(
        slug='tpp', cat='production', catLabel='Производство',
        logo='assets/img/logos/TPP-logo 26.webp', image=10,
        shortLoc='Ташкент, ул. Олтин-Топган',
        breadcrumb='Toshkent Plast Polimer',
        h1_l1='Toshkent', h1_l2='Plast Polimer',
        overviewH1='Полимерная упаковка и', overviewH2='пластмассовые изделия',
        relatedHeading='производственные активы',
        banner='assets/img/projects/plast-polimer/banner.webp',
        gallery=[
            'assets/img/projects/plast-polimer/1.webp',
            'assets/img/projects/plast-polimer/2.webp',
            'assets/img/projects/plast-polimer/3.webp',
            'assets/img/projects/plast-polimer/4.webp',
            'assets/img/projects/plast-polimer/5.webp',
            'assets/img/projects/plast-polimer/6.webp',
        ],
    ),
    '9. BA Texnopark': dict(
        slug='ba-texnopark', cat='production', catLabel='Производство',
        logo='assets/img/logos/BA Texnopark-removebg-preview 26.webp', image=11,
        shortLoc='Сырдарьинская обл., Хакикат',
        siteLabel='wangdagroup.uz', siteUrl='https://wangdagroup.uz/',
        breadcrumb='BA Texnopark',
        h1_l1='BA Texnopark', h1_l2='(Wang Da Group)',
        overviewH1='Малая промышленная зона', overviewH2='для бизнеса',
        relatedHeading='производственные активы',
        banner='assets/img/projects/other/other (11).webp',
        gallery=[
            'assets/img/projects/other/other (12).webp',
            'assets/img/projects/other/other (13).webp',
            'assets/img/projects/other/other (14).webp',
            'assets/img/projects/other/other (15).webp',
        ],
    ),
    '10. Awiner AGRО': dict(
        slug='awiner-agro', cat='agro', catLabel='Агрохимия',
        logo='assets/img/logos/AWINER AGRO logo 26.webp', image=12,
        shortLoc='Ташкентская обл., Ахангаранский р-н',
        breadcrumb='Awiner AGRO',
        h1_l1='Awiner AGRO', h1_l2='Агрохимия и защита растений',
        overviewH1='Химия для', overviewH2='современного сельского хозяйства',
        relatedHeading='агроактивы',
        banner='assets/img/projects/agro/banner.webp',
        gallery=[
            'assets/img/projects/agro/1.webp',
            'assets/img/projects/agro/2.webp',
            'assets/img/projects/agro/3.webp',
        ],
    ),
    '11. Sirdarya Metal Construction': dict(
        slug='sirdarya-metal-construction', cat='production', catLabel='Производство',
        logo='assets/img/logos/Sirdarya Metal Construction 26.webp', image=7,
        shortLoc='Сырдарьинская обл., Хакикат',
        breadcrumb='Sirdarya Metal Construction',
        h1_l1='Sirdarya Metal', h1_l2='Construction',
        overviewH1='Металлические конструкции', overviewH2='для промышленных объектов',
        relatedHeading='производственные активы',
        banner='assets/img/projects/other/other (16).webp',
        gallery=[
            'assets/img/projects/other/other (17).webp',
            'assets/img/projects/other/other (18).webp',
            'assets/img/projects/other/other (19).webp',
            'assets/img/projects/other/other (20).webp',
        ],
    ),
    '12. Shengli Steel': dict(
        slug='shengli-steel', cat='metallurgy', catLabel='Металлургия',
        logo='assets/img/logos/Shengli Steel 26.webp', image=9,
        shortLoc='Ташкентская обл., г. Бекабад',
        breadcrumb='Shengli Steel',
        h1_l1='Shengli Steel', h1_l2='Переработка металлургического шлака',
        overviewH1='Вторичная переработка', overviewH2='металлургического шлака',
        relatedHeading='металлургические активы',
        banner='assets/img/projects/other/other (21).webp',
        gallery=[
            'assets/img/projects/other/other (22).webp',
            'assets/img/projects/other/other (23).webp',
            'assets/img/projects/other/other (24).webp',
            'assets/img/projects/other/other (25).webp',
        ],
    ),
    '13. Tai Chang Group': dict(
        slug='tai-chang-group', cat='metallurgy', catLabel='Металлургия',
        logo='assets/img/logos/Ti Chang Group 26-3.webp', image=11,
        shortLoc='Ташкентская обл., Ахангаранский р-н',
        breadcrumb='Tai Chang Group',
        h1_l1='Tai Chang Group', h1_l2='Высоколегированный металлопрокат',
        overviewH1='Крупногабаритный', overviewH2='металлопрокат',
        relatedHeading='металлургические активы',
        banner='assets/img/projects/other/other (26).webp',
        gallery=[
            'assets/img/projects/other/other (27).webp',
            'assets/img/projects/other/other (28).webp',
        ],
    ),
}

# ---------- 3. Cleaning helpers ----------
EMOJI_RE = re.compile(
    r'[\U0001F1E6-\U0001F1FF\U0001F300-\U0001FAFF\U00002600-\U000027BF]+',
    re.UNICODE,
)

def clean(s):
    if not s:
        return s
    s = EMOJI_RE.sub('', s).strip()
    s = re.sub(r'\s+', ' ', s).rstrip('.').strip()
    return s

# ---------- 4. Products parsing ----------
def parse_products(text):
    """Excel "Products" column uses two formats — either numbered
    (01\\nTitle\\nDescription\\n\\n02\\n…) or a plain bullet list."""
    if not text:
        return []
    lines = [l.rstrip() for l in text.split('\n')]
    starts = [(i, l.strip()) for i, l in enumerate(lines) if re.match(r'^0[1-9]$', l.strip())]
    if len(starts) >= 2:
        groups = []
        for idx, (i, num) in enumerate(starts):
            end = starts[idx + 1][0] if idx + 1 < len(starts) else len(lines)
            chunk = [l.strip() for l in lines[i + 1:end] if l.strip()]
            if not chunk:
                continue
            title = chunk[0]
            body = ' '.join(chunk[1:]).strip()
            groups.append({'num': num, 'title': title, 'text': body})
        return groups[:4]
    items = [l.strip() for l in lines if l.strip()]
    return [{'num': f'0{i + 1}', 'title': it, 'text': ''} for i, it in enumerate(items[:4])]

# ---------- 5. Build per-project records ----------
def build_projects(raw):
    projects = []
    for sheet, base in raw.items():
        if sheet not in META:
            continue
        m = META[sheet]
        p = dict(base)
        p.update(m)
        for f in ('name', 'location', 'desc'):
            if f in base:
                p[f] = clean(base[f])
        p['products'] = parse_products(base.get('products', ''))
        p['turnover_clean']  = base.get('turnover') or '—'
        p['employees_clean'] = str(base.get('employees', '')).strip() or '—'
        p['capacity_clean']  = base.get('capacity') or '—'
        p['year_clean']      = base.get('year') or '—'
        p['email_clean']     = (base.get('email')  or '').strip() or None
        p['phone_clean']     = (base.get('phone')  or '').strip() or None
        site = base.get('site') or ''
        if site and site != '-' and not m.get('siteUrl'):
            p['siteUrl']   = site
            p['siteLabel'] = re.sub(r'^https?://(www\.)?', '', site).rstrip('/')
        p['lead']       = (base.get('desc') or '').replace('\n', ' ').strip()
        about = base.get('about_full') or ''
        p['about_paras'] = [
            pp.replace('\n', ' ').strip()
            for pp in re.split(r'\n\s*\n', about) if pp.strip()
        ]
        projects.append(p)
    return projects

# ---------- 6. Renderers (per-project HTML fragments) ----------
def short_desc(p):
    d = (p.get('desc') or '').replace('\n', ' ').strip()
    return d[:160] + ('…' if len(d) > 163 else '')

def card_image(p):
    """Use the per-project banner image when available (overlaid with the
    existing dark gradient so card text stays readable); otherwise fall back
    to the colour-tuned `.proj-card__image--N` background defined in
    projects.css."""
    if p.get('banner'):
        return (f'<div class="proj-card__image" '
                f'style="background-image: linear-gradient(135deg,rgba(14,28,66,0.5),rgba(8,18,46,0.65)), '
                f"url('{p['banner']}'); background-size: cover; background-position: center;\"></div>")
    return f'<div class="proj-card__image proj-card__image--{p["image"]}"></div>'

def render_related(curr, projects):
    same_cat = [p for p in projects if p['cat'] == curr['cat'] and p['slug'] != curr['slug']]
    others   = [p for p in projects if p['cat'] != curr['cat']]
    picks    = (same_cat + others)[:3]
    cards = []
    for p in picks:
        emp = p['employees_clean']
        emp_str = f"{emp} сотр." if emp != '—' else '—'
        cards.append(f'''        <a href="project-{p['slug']}.html" class="proj-card">
          <div class="proj-card__image-wrap">
            {card_image(p)}
          </div>
          <div class="proj-card__body">
            <span class="proj-card__cat">{p['catLabel']}</span>
            <h3 class="proj-card__title">{p['name']}</h3>
            <p class="proj-card__desc">{short_desc(p)}</p>
            <div class="proj-card__meta">
              <div class="proj-card__meta-row">
                <span class="proj-card__meta-key">{p['shortLoc']}</span>
                <span class="proj-card__meta-val">{emp_str}</span>
              </div>
            </div>
          </div>
        </a>''')
    return '\n'.join(cards)

def render_features(prods):
    if not prods:
        return '            <p class="pi-overview__para">Подробная номенклатура продукции уточняется.</p>'
    blocks = []
    for prod in prods:
        text_html = (
            f'                <p class="pi-feature__text">{prod["text"]}</p>'
            if prod['text'] else ''
        )
        blocks.append(f'''            <div class="pi-feature">
              <span class="pi-feature__num">{prod['num']}</span>
              <div>
                <h4 class="pi-feature__title">{prod['title']}</h4>
{text_html}
              </div>
            </div>'''.replace('\n\n', '\n'))
    return '\n'.join(blocks)

def render_stats(p):
    items = [
        (p['turnover_clean'],  'Годовой оборот'),
        (p['employees_clean'], 'Сотрудников'),
        (p['capacity_clean'],  'Мощность в год'),
        (str(p['year_clean']), 'Год запуска'),
    ]
    return '\n'.join(
        f'''        <div class="pi-stats__item">
          <span class="pi-stats__num">{num}</span>
          <span class="pi-stats__lbl">{lbl}</span>
        </div>'''
        for num, lbl in items
    )

def render_paragraphs(paras):
    if not paras:
        return ''
    return '\n\n'.join(
        f'          <p class="pi-overview__para">\n            {pp}\n          </p>'
        for pp in paras[:3]
    )

# ---------- 7. Main per-project file generator ----------
def make_project_file(template, p, projects):
    out = template
    # title
    out = re.sub(
        r'<title>[^<]+— B\.A\. Holding</title>',
        f'<title>{p["name"]} — B.A. Holding</title>',
        out, count=1,
    )
    # breadcrumb
    out = re.sub(
        r'<span class="breadcrumb__current">[^<]+</span>',
        f'<span class="breadcrumb__current">{p["breadcrumb"]}</span>',
        out, count=1,
    )
    # hero category
    out = re.sub(
        r'<span class="pi-hero__cat">[^<]*</span>',
        f'<span class="pi-hero__cat">{p["catLabel"]}</span>',
        out, count=1,
    )
    # hero H1
    out = re.sub(
        r'<h1 class="page-hero__title">[^<]*<br>[^<]*</h1>',
        f'<h1 class="page-hero__title">{p["h1_l1"]}<br>{p["h1_l2"]}</h1>',
        out, count=1,
    )
    # hero lead
    out = re.sub(
        r'<p class="pi-hero__lead">\s*[^<]+\s*</p>',
        f'<p class="pi-hero__lead">\n        {p["lead"]}\n      </p>',
        out, count=1, flags=re.S,
    )
    # hero background image (banner) — only if the project has its own banner
    if p.get('banner'):
        out = re.sub(
            r'(<div class="page-hero__bg">\s*<img src=")[^"]+(")',
            lambda m: m.group(1) + p['banner'] + m.group(2),
            out, count=1, flags=re.S,
        )
    # gallery slider — replace 4 default slides with project's gallery list
    if p.get('gallery'):
        slides_html = '\n'.join(
            f'''            <div class="pi-gallery__slide">
              <div class="pi-gallery__img" style="background-image:url('{src}')"></div>
            </div>'''
            for src in p['gallery']
        )
        out = re.sub(
            r'(<div class="pi-gallery__track" data-pi-track>)(.*?)(</div>\s*</div>\s*<button)',
            lambda m: m.group(1) + '\n' + slides_html + '\n          ' + m.group(3),
            out, count=1, flags=re.S,
        )
    # stats grid
    out = re.sub(
        r'(<div class="pi-stats__grid">)(.*?)(</div>\s*</div>\s*</section>)',
        lambda m: m.group(1) + '\n' + render_stats(p) + '\n      ' + m.group(3),
        out, count=1, flags=re.S,
    )
    # overview heading
    out = re.sub(
        r'<h2 class="section-heading">[^<]*<br><span class="fw-semibold">[^<]*</span></h2>',
        f'<h2 class="section-heading">{p["overviewH1"]}<br><span class="fw-semibold">{p["overviewH2"]}</span></h2>',
        out, count=1,
    )
    # overview paragraphs
    out = re.sub(
        r'(<h2 class="section-heading">[^<]*<br><span class="fw-semibold">[^<]*</span></h2>)\s*(<p class="pi-overview__para">.*?</p>\s*){1,3}',
        lambda m: m.group(1) + '\n\n' + render_paragraphs(p['about_paras']) + '\n\n          ',
        out, count=1, flags=re.S,
    )
    # features
    out = re.sub(
        r'(<div class="pi-overview__features">)(.*?)(</div>\s*</div>\s*<!-- Sidebar logo card -->)',
        lambda m: m.group(1) + '\n' + render_features(p['products']) + '\n          ' + m.group(3),
        out, count=1, flags=re.S,
    )
    # sidebar logo (replaces former info-card)
    out = re.sub(
        r'<img class="pi-logo-card__img" src="[^"]+" alt="[^"]*"',
        f'<img class="pi-logo-card__img" src="{p["logo"]}" alt="{p["name"]}"',
        out, count=1,
    )
    # related heading
    out = re.sub(
        r'<h2 class="section-heading">Другие <span class="fw-semibold">[^<]*</span></h2>',
        f'<h2 class="section-heading">Другие <span class="fw-semibold">{p["relatedHeading"]}</span></h2>',
        out, count=1,
    )
    # related grid
    out = re.sub(
        r'(<div class="pi-related__grid">)(.*?)(</div>\s*</div>\s*</section>\s*<!-- ===== CTA ===== -->)',
        lambda m: m.group(1) + '\n' + render_related(p, projects) + '\n      ' + m.group(3),
        out, count=1, flags=re.S,
    )
    # category in CTAs / "Все проекты направления"
    out = out.replace('projects.html?cat=metallurgy', 'projects.html?cat=' + p['cat'])
    return out

# ---------- 8. projects.html cards renderer ----------
def render_projects_card(p):
    emp = p['employees_clean']
    emp_html = (
        f'{emp} <span data-i18n="projects.meta.employees_short">сотр.</span>'
        if emp != '—' else '—'
    )
    second = ''
    if p['turnover_clean'] != '—':
        second = (
            '\n              <div class="proj-card__meta-row">'
            '\n                <span class="proj-card__meta-key" data-i18n="projects.meta.turnover">Оборот</span>'
            f'\n                <span class="proj-card__meta-val">{p["turnover_clean"]}</span>'
            '\n              </div>'
        )
    elif p['capacity_clean'] != '—':
        second = (
            '\n              <div class="proj-card__meta-row">'
            '\n                <span class="proj-card__meta-key" data-i18n="projects.meta.capacity">Мощность</span>'
            f'\n                <span class="proj-card__meta-val">{p["capacity_clean"]}</span>'
            '\n              </div>'
        )
    return f'''        <a class="proj-card" href="project-{p['slug']}.html" data-cat="{p['cat']}" data-status="done">
          <div class="proj-card__image-wrap">
            {card_image(p)}
          </div>
          <div class="proj-card__body">
            <span class="proj-card__cat">{p['catLabel']}</span>
            <h3 class="proj-card__title">{p['name']}</h3>
            <p class="proj-card__desc">{short_desc(p)}</p>
            <div class="proj-card__meta">
              <div class="proj-card__meta-row">
                <span class="proj-card__meta-key">{p['shortLoc']}</span>
                <span class="proj-card__meta-val">{emp_html}</span>
              </div>{second}
            </div>
          </div>
        </a>'''

def update_projects_html(projects):
    cards = '\n\n'.join(render_projects_card(p) for p in projects)
    src = PROJECTS_HTML.read_text(encoding='utf-8')
    pat = re.compile(
        r'(<div class="proj-grid" id="projGrid">)(.*?)(\s*</div>\s*<p class="portfolio__count")',
        re.S,
    )
    src, n = pat.subn(lambda m: m.group(1) + '\n\n' + cards + '\n\n      ' + m.group(3), src)
    if n != 1:
        raise SystemExit('Could not locate <div class="proj-grid" id="projGrid"> block in projects.html')
    src = re.sub(
        r'<p class="portfolio__count" id="projCount">[^<]*</p>',
        f'<p class="portfolio__count" id="projCount">Всего {len(projects)} предприятий</p>',
        src, count=1,
    )
    PROJECTS_HTML.write_text(src, encoding='utf-8', newline='\n')

# ---------- 9. Entry point ----------
def main():
    raw = load_excel()
    projects = build_projects(raw)
    if not projects:
        raise SystemExit('No projects loaded — check META keys match Excel sheet names.')

    template = TEMPLATE.read_text(encoding='utf-8')

    written = []
    for p in projects:
        out_path = ROOT / f'project-{p["slug"]}.html'
        out_path.write_text(make_project_file(template, p, projects), encoding='utf-8', newline='\n')
        written.append(out_path.name)

    update_projects_html(projects)

    print(f'Wrote {len(written)} project pages:')
    for w in written:
        print(' ', w)
    print(f'Updated projects.html ({len(projects)} cards).')

if __name__ == '__main__':
    main()
